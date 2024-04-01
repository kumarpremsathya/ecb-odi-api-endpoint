import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import openpyxl
import xlrd
import mysql.connector
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet.header_footer")
xpath_2004 = [
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[2]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[4]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[6]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[8]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[10]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[12]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[14]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[16]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[18]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[20]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[22]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[24]/td[3]/a"
    ]


xpath_2005 = [
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[2]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[4]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[6]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[8]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[10]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[12]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[14]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[16]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[18]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[20]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[22]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[24]/td[2]/a"
    ]

xpath_2006 = [
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[2]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[4]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[6]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[8]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[10]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[12]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[14]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[16]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[18]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[20]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[22]/td[3]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[24]/td[3]/a"
    ]

def read_xls(filename):
            workbook = xlrd.open_workbook(filename)
            sheet = workbook.sheet_by_index(0)
            return sheet

def read_xlsx(filename):
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            return sheet

def download_files_for_year(year, months_xpath):
    i = 2
    download_base_directory = fr"C:\Users\magudapathy.7409\Desktop\RBI_excel_2\magu"
    download_directory = os.path.join(download_base_directory, str(year))

    if not os.path.exists(download_directory):
        os.makedirs(download_directory)

    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_directory,
        "download.default_content_setting_value": 2,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True
    }
    chrome_options.add_experimental_option("prefs", prefs)

    try:
        browser = webdriver.Chrome(options=chrome_options)

        browser.get("https://rbi.org.in/Scripts/ECBView.aspx")
        time.sleep(5)

        archive = browser.find_element(By.XPATH, '//*[@id="divArchiveMain"]')
        archive.click()
        time.sleep(5)

        next_xpath = browser.find_element(By.XPATH, f'//*[@id="{year}"]')
        next_xpath.click()
        time.sleep(5)

        specific_xpaths = globals()[f'xpath_{year}'] if year in ["2004", "2005", "2006"] else months_xpath

        failed_downloads = []

        for xpath in specific_xpaths:
            time.sleep(7)
            month_xpath = browser.find_element(By.XPATH, xpath)

            original_file_name = month_xpath.get_attribute("href").split("/")[-1]
            original_file_name = original_file_name.replace("%20", " ")
            print("original file name as in the website is ", original_file_name)

            downloaded_files = os.listdir(download_directory)
            print(f"the files which are in the {year} folder", downloaded_files)

            rename_xpath = browser.find_element(By.XPATH, f"//*[@id='annual']/div[2]/table[2]/tbody/tr[{i}]/td[1]/a")
            i = i + 2

            original_extension=os.path.splitext(original_file_name)[1]
            print("original_extension",original_extension)


            if year in ["2004", "2005"]:
                new_file_name = rename_xpath.text.split("-")[-1].strip()
                print(new_file_name)
            else:
                new_file_name = rename_xpath.text.split("of")[-1].strip()
                print(new_file_name)

            
            new_file_name_ex=new_file_name + original_extension
            print(new_file_name_ex)


            if new_file_name_ex in downloaded_files or original_file_name in downloaded_files:
                print(f"{year}/{new_file_name_ex} or {original_file_name}- File already exists. Skipping download...")
                
            else:
                try:
                    time.sleep(10)
                    month_xpath.click()
                    time.sleep(30)
                    current_file_name = os.path.join(download_directory, original_file_name)
                

                    if original_extension == ".xls":
                        rename_file_path = os.path.join(download_directory, new_file_name_ex)
                        time.sleep(15)
                        os.rename(current_file_name, rename_file_path )
                        print(f"{year}/{new_file_name_ex} - downloaded...")

                    elif original_extension == ".xlsx":
                        rename_file_path = os.path.join(download_directory, new_file_name_ex)
                        time.sleep(15)
                        os.rename(current_file_name, rename_file_path)
                        print(f"{year}/{new_file_name_ex} - downloaded...")

                    else:    
                        original_extension == ".pdf"
                        rename_file_path = os.path.join(download_directory, new_file_name_ex)
                        time.sleep(15)
                        os.rename(current_file_name, rename_file_path)
                        print(f"{year}/{new_file_name_ex} - downloaded...")
        
                except Exception as e:
                    print(f"An unexpected error occurred: {str(e)}")
                    failed_downloads.append((month_xpath, new_file_name))

        # Retry failed downloads
        for xpath, new_file_name_ex in failed_downloads:
            try:
                time.sleep(10)
                xpath.click()
                time.sleep(30)

                if original_extension == ".xls":
                    rename_file_path = os.path.join(download_directory, new_file_name_ex)
                    time.sleep(15)
                    os.rename(current_file_name, rename_file_path )
                    print(f"{year}/{new_file_name_ex} - downloaded...")

                elif original_extension == ".xlsx":
                    rename_file_path = os.path.join(download_directory, new_file_name_ex)
                    time.sleep(15)
                    os.rename(current_file_name, rename_file_path)
                    print(f"{year}/{new_file_name_ex} - downloaded...")

                else:    
                    original_extension == ".pdf"
                    rename_file_path = os.path.join(download_directory, new_file_name_ex)
                    time.sleep(15)
                    os.rename(current_file_name, rename_file_path)
                    print(f"{year}/{new_file_name_ex} - downloaded...")
                
            except Exception as e:
                print(f"An unexpected error occurred during retry: {str(e)}")

    except TimeoutException:
        print(f"Timeout exception occurred for {year}. Please check the internet connection and try again.")
    except NoSuchElementException:
        print(f"Element not found exception occurred for {year}. Please check if the webpage structure has changed.")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")

    finally:
        if 'browser' in locals() and browser is not None:
            browser.quit()

# Define the list of years
years = ["2004", "2005", "2006","2007", "2008", "2009", "2010", "2011", "2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020", "2021", "2022", "2023"]


def extract_and_insert_data(filename, cursor):
            _, file_extension = os.path.splitext(filename)

            if file_extension.lower() == '.xls':
                try:
                    sheet = read_xls(filename)
                except xlrd.biffh.XLRDError:
                    filename = filename.replace('.xls', '.xlsx')
                    sheet = read_xlsx(filename)
            elif file_extension.lower() == '.xlsx':
                sheet = read_xlsx(filename)
            else:
                print(f"Unsupported file format: {file_extension}")
                return

            year_counter = 0
            automated_counter = 0
            approved_counter = 0
            automated = False
            approved = False
            new_list = []

            if isinstance(sheet, xlrd.sheet.Sheet):  # Logic for xls files
                month,year="",""
                for row_index in range(sheet.nrows):
                    e_list = [str(sheet.cell_value(row_index, col)) for col in range(sheet.ncols) if sheet.cell_value(row_index, col) is not None]

                    while "" in e_list:
                        e_list.remove("")

                    if not e_list:
                        continue

                    for names in e_list:
                        try:
                            if "month" in names.lower() and year_counter == 0:
                                month_year = names.split("for the month of")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if "-" in names.lower() and year_counter == 0:
                                month_year = names.split("-")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if year_counter == 0:
                                month_year = filename.split('\\')[-1].split('.')[0]
                                month, year = month_year.split()
                                year_counter += 1
                        except Exception as e:
                            print(e)

                        if "automatic route" in names.lower() and automated_counter == 0:
                            automated = True
                            approved = False
                            automated_counter += 1
                        if "approval route" in names.lower() and approved_counter == 0:
                            approved = True
                            automated = False
                            approved_counter += 1

                    if automated:
                        new_list = e_list + [month, year, "Automatic"]
                    elif approved:
                        new_list = e_list + [month, year, "Approval"]

                    last_list = [name.replace(".0", "") if isinstance(name, str) and name.endswith(".0") else name for name in new_list]

                    # if last_list and "Approval" in new_list and len(last_list) >= 9:
                    #     insert_data_into_database(cursor, last_list, "Approval")

                    # if last_list and "Automatic" in new_list and len(last_list) >= 9:
                    #     insert_data_into_database(cursor, last_list, "Automatic")
                    
                    if last_list and "Approval" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Approval")

                    if last_list and "Automatic" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Automatic")


            elif isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                month,year="",""# Logic for xlsx files
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                    e_list = [str(cell) for cell in row if cell is not None]

                    while "" in e_list:
                        e_list.remove("")

                    if not e_list:
                        continue

                    for names in e_list:
                        try:
                            if "month" in names.lower() and year_counter == 0:
                                month_year = names.split("for the month of")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if "-" in names.lower() and year_counter == 0:
                                month_year = names.split("-")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if year_counter == 0:
                                month_year = filename.split('\\')[-1].split('.')[0]
                                month, year = month_year.split()
                                year_counter += 1
                        except Exception as e:
                            print(e)

                        if "automatic route" in names.lower() and automated_counter == 0:
                            automated = True
                            approved = False
                            automated_counter += 1
                        if "approval route" in names.lower() and approved_counter == 0:
                            approved = True
                            automated = False
                            approved_counter += 1

                    if automated:
                        new_list = e_list + [month, year, "Automatic"]
                    elif approved:
                        new_list = e_list + [month, year, "Approval"]

                    last_list = [name.replace(".0", "") if isinstance(name, str) and name.endswith(".0") else name for name in new_list]

                    if last_list and "Approval" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Approval")

                    if last_list and "Automatic" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Automatic")

def insert_data_into_database(cursor, data_list, route):
            month, year = data_list[-3], data_list[-2]

            data_dict = {
                'ECB_FCCB': data_list[1],
                'Borrower': data_list[2],
                'Equivalent_Amount_in_USD': data_list[3],
                'Purpose': data_list[4],
                'Maturity_Period': data_list[5],
                'Month': month,  # Corrected: Assigning the correct value to the 'Month' column
                'Year': year,    # Corrected: Assigning the correct value to the 'Year' column
                'Route': route,
            }

            cursor.execute(
                f"INSERT INTO new_2 ({', '.join(data_dict.keys())}) VALUES ({', '.join(['%s']*len(data_dict))})",
                tuple(data_dict.values())
            )

def insert_data_into_database_11(cursor, data_list, route):
            month, year = data_list[-3], data_list[-2]

            data_dict = {
                'ECB_FCCB': data_list[1],
                'Borrower': data_list[2],
                'Economic_sector_of_borrower': data_list[3] if data_list[3] else None,
                'Equivalent_Amount_in_USD': data_list[4],
                'Purpose': data_list[5],
                'Maturity_Period': data_list[6],
                'Lender_Category': data_list[7] if data_list[7] else None,
                'Month': month,  # Corrected: Assigning the correct value to the 'Month' column
                'Year': year,    # Corrected: Assigning the correct value to the 'Year' column
                'Route': route,
            }

            cursor.execute(
                f"INSERT INTO new_2 ({', '.join(data_dict.keys())}) VALUES ({', '.join(['%s']*len(data_dict))})",
                tuple(data_dict.values())
            )

def insert_data_into_database_49(cursor, data_list, route):
            month, year = data_list[-3], data_list[-2]
            

            data_dict = {
                'ECB_FCCB': data_list[0],
                'Borrower': data_list[1],
                'Equivalent_Amount_in_USD': data_list[2],
                'Purpose': data_list[3],
                'Maturity_Period': data_list[4],
                'Month': month,  # Corrected: Assigning the correct value to the 'Month' column
                'Year': year,    # Corrected: Assigning the correct value to the 'Year' column
                'Route': route,
            }

            cursor.execute(
                f"INSERT INTO new_2 ({', '.join(data_dict.keys())}) VALUES ({', '.join(['%s']*len(data_dict))})",
                tuple(data_dict.values())
            )

def del2(cursor):
    cursor.execute("DELETE FROM new_2 WHERE Equivalent_Amount_in_USD='Equivalent Amount in USD ' or Borrower='Equivalent Amount in USD' or ECB_FCCB='Borrower'or Borrower='Borrower' ")

# Main function
def main():
    # Define the list of years
    years = ["2004", "2005", "2006","2007", "2008", "2009", "2010", "2011", "2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020", "2021", "2022", "2023"]

    # Set up database connection
    db_config = {
        'host': 'localhost',
        'user': 'root',
        'password': 'root',
        'database': 'rbi_main1',
        "auth_plugin": "mysql_native_password"
    }
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    # Create the table if it doesn't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS new_2(
            ECB_FCCB VARCHAR(255),
            Borrower VARCHAR(255),
            Economic_sector_of_borrower VARCHAR(255),
            Equivalent_Amount_in_USD MEDIUMTEXT,
            Purpose VARCHAR(255),   
            Maturity_Period VARCHAR(255),
            Lender_Category VARCHAR(255),
            Month VARCHAR(255),
            Year VARCHAR(255),
            Route VARCHAR(255),
            scraped_on TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Set the base path for extracted files
    base_path = r"C:\Users\magudapathy.7409\Desktop\RBI_excel_2\magu"

    # Define the months_xpath variable
    months_xpath = [
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[2]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[4]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[6]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[8]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[10]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[12]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[14]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[16]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[18]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[20]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[22]/td[2]/a",
        "//*[@id='annual']/div[2]/table[2]/tbody/tr[24]/td[2]/a"
    ]


    # Iterate over years and download files
    for year in years:
        download_files_for_year(year, months_xpath)

    # Iterate over folders and files for extraction
    folders = os.listdir(base_path)
    for folder in folders:
        folder_path = os.path.join(base_path, folder)
        year_files = os.listdir(folder_path)

        for filename in year_files:
            file_path = os.path.join(folder_path, filename)

            # Handle different file extensions
            if not os.path.isfile(file_path) and filename.lower().endswith('.xls'):
                file_path = file_path.replace('.xls', '.xlsx')


            extract_and_insert_data(file_path, cursor)

    # Commit the changes and close the connection
    conn.commit()
    del2(cursor)
    conn.commit()
    conn.close()


# Run the main function
if __name__ == "__main__":
    main()
