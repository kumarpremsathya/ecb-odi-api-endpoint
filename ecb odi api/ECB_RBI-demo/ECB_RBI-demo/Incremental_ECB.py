# final version of incremental for ecb
import mysql.connector
import os
import time
import openpyxl
import xlrd
from selenium import webdriver
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import traceback
import sys

actual_months = ['November', 'December', 'June', 'September', 'February', 'July', 'August', 'October', 'January', 'April', 'May', 'March']

month_xpath = {"January": "(//tbody//tr//a[contains(text(),'January')]//ancestor::tr//td//a)[2]",
               "February": "(//tbody//tr//a[contains(text(),'February')]//ancestor::tr//td//a)[2]",
               "March": "(//tbody//tr//a[contains(text(),'March')]//ancestor::tr//td//a)[2]",
               "April": "(//tbody//tr//a[contains(text(),'April')]//ancestor::tr//td//a)[2]",
               "May": "(//tbody//tr//a[contains(text(),'May')]//ancestor::tr//td//a)[2]",
               "June": "(//tbody//tr//a[contains(text(),'June')]//ancestor::tr//td//a)[2]",
               "July": "(//tbody//tr//a[contains(text(),'July')]//ancestor::tr//td//a)[2]",
               "August": "(//tbody//tr//a[contains(text(),'August')]//ancestor::tr//td//a)[2]",
               "September": "(//tbody//tr//a[contains(text(),'September')]//ancestor::tr//td//a)[2]",
               "October": "(//tbody//tr//a[contains(text(),'October')]//ancestor::tr//td//a)[2]",
               "November": "(//tbody//tr//a[contains(text(),'November')]//ancestor::tr//td//a)[2]",
               "December": "(//tbody//tr//a[contains(text(),'December')]//ancestor::tr//td//a)[2]"}

filename_xpath = {"January": "(//tbody//tr//a[contains(text(),'January')]//ancestor::tr//td//a)[1]",
                  "February": "(//tbody//tr//a[contains(text(),'February')]//ancestor::tr//td//a)[1]",
                  "March": "(//tbody//tr//a[contains(text(),'March')]//ancestor::tr//td//a)[1]",
                  "April": "(//tbody//tr//a[contains(text(),'April')]//ancestor::tr//td//a)[1]",
                  "May": "(//tbody//tr//a[contains(text(),'May')]//ancestor::tr//td//a)[1]",
                  "June": "(//tbody//tr//a[contains(text(),'June')]//ancestor::tr//td//a)[1]",
                  "July": "(//tbody//tr//a[contains(text(),'July')]//ancestor::tr//td//a)[1]",
                  "August": "(//tbody//tr//a[contains(text(),'August')]//ancestor::tr//td//a)[1]",
                  "September": "(//tbody//tr//a[contains(text(),'September')]//ancestor::tr//td//a)[1]",
                  "October": "(//tbody//tr//a[contains(text(),'October')]//ancestor::tr//td//a)[1]",
                  "November": "(//tbody//tr//a[contains(text(),'November')]//ancestor::tr//td//a)[1]",
                  "December": "(//tbody//tr//a[contains(text(),'December')]//ancestor::tr//td//a)[1]"}



# Define your database configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'ecb_rbi',
    'auth_plugin': 'mysql_native_password'
}


conn = mysql.connector.connect(**db_config)
cursor = conn.cursor()


def read_xls(filename):
            workbook = xlrd.open_workbook(filename)
            sheet = workbook.sheet_by_index(0)
            return sheet

def read_xlsx(filename):
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            return sheet







def extract_and_insert_data(filename, cursor):
            _, file_extension = os.path.splitext(filename)

            if file_extension.lower() == '.xls':
                try:
                    sheet = read_xls(filename)
                except xlrd.biffh.XLRDError:
                    filename = filename.replace('.xls', '.xlsx')
                    sheet = read_xlsx(filename)
            elif file_extension.lower() == '.xlsx':
                sheet = read_xlsx(filename)
            else:
                print(f"Unsupported file format: {file_extension}")
                return

            year_counter = 0
            automated_counter = 0
            approved_counter = 0
            automated = False
            approved = False
            new_list = []

            if isinstance(sheet, xlrd.sheet.Sheet):  # Logic for xls files
                month,year="",""
                for row_index in range(sheet.nrows):
                    e_list = [str(sheet.cell_value(row_index, col)) for col in range(sheet.ncols) if sheet.cell_value(row_index, col) is not None]

                    while "" in e_list:
                        e_list.remove("")

                    if not e_list:
                        continue

                    for names in e_list:
                        try:
                            if "month" in names.lower() and year_counter == 0:
                                month_year = names.split("for the month of")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if "-" in names.lower() and year_counter == 0:
                                month_year = names.split("-")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if year_counter == 0:
                                month_year = filename.split('\\')[-1].split('.')[0]
                                month, year = month_year.split()
                                year_counter += 1
                        except Exception as e:
                            print(e)

                        if "automatic route" in names.lower() and automated_counter == 0:
                            automated = True
                            approved = False
                            automated_counter += 1
                        if "approval route" in names.lower() and approved_counter == 0:
                            approved = True
                            automated = False
                            approved_counter += 1

                    if automated:
                        new_list = e_list + [month, year, "Automatic"]
                    elif approved:
                        new_list = e_list + [month, year, "Approval"]

                    last_list = [name.replace(".0", "") if isinstance(name, str) and name.endswith(".0") else name for name in new_list]

                    # if last_list and "Approval" in new_list and len(last_list) >= 9:
                    #     insert_data_into_database(cursor, last_list, "Approval")

                    # if last_list and "Automatic" in new_list and len(last_list) >= 9:
                    #     insert_data_into_database(cursor, last_list, "Automatic")
                    
                    if last_list and "Approval" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Approval")

                    if last_list and "Automatic" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Automatic")


            elif isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                month,year="",""# Logic for xlsx files
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                    e_list = [str(cell) for cell in row if cell is not None]

                    while "" in e_list:
                        e_list.remove("")

                    if not e_list:
                        continue

                    for names in e_list:
                        try:
                            if "month" in names.lower() and year_counter == 0:
                                month_year = names.split("for the month of")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if "-" in names.lower() and year_counter == 0:
                                month_year = names.split("-")[1].strip()
                                month, year = month_year.split()
                                year_counter += 1
                            if year_counter == 0:
                                month_year = filename.split('\\')[-1].split('.')[0]
                                month, year = month_year.split()
                                year_counter += 1
                        except Exception as e:
                            print(e)

                        if "automatic route" in names.lower() and automated_counter == 0:
                            automated = True
                            approved = False
                            automated_counter += 1
                        if "approval route" in names.lower() and approved_counter == 0:
                            approved = True
                            automated = False
                            approved_counter += 1

                    if automated:
                        new_list = e_list + [month, year, "Automatic"]
                    elif approved:
                        new_list = e_list + [month, year, "Approval"]

                    last_list = [name.replace(".0", "") if isinstance(name, str) and name.endswith(".0") else name for name in new_list]

                    if last_list and "Approval" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Approval")
                    elif last_list and "Approval" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Approval")

                    if last_list and "Automatic" in new_list and len(last_list) >= 10:
                        insert_data_into_database_11(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 9:
                        insert_data_into_database(cursor, last_list, "Automatic")
                    elif last_list and "Automatic" in new_list and len(last_list) >= 8:
                        insert_data_into_database_49(cursor, last_list, "Automatic")

def insert_data_into_database(cursor, data_list, route):
            month, year = data_list[-3], data_list[-2]

            data_dict = {
                'ECB_FCCB': data_list[1],
                'Borrower': data_list[2],
                'Equivalent_Amount_in_USD': data_list[3],
                'Purpose': data_list[4],
                'Maturity_Period': data_list[5],
                'Month': month,  # Corrected: Assigning the correct value to the 'Month' column
                'Year': year,    # Corrected: Assigning the correct value to the 'Year' column
                'Route': route,
            }

            cursor.execute(
                f"INSERT INTO 2005_2023_1 ({', '.join(data_dict.keys())}) VALUES ({', '.join(['%s']*len(data_dict))})",
                tuple(data_dict.values())
            )

def insert_data_into_database_11(cursor, data_list, route):
            month, year = data_list[-3], data_list[-2]

            data_dict = {
                'ECB_FCCB': data_list[1],
                'Borrower': data_list[2],
                'Economic_sector_of_borrower': data_list[3] if data_list[3] else None,
                'Equivalent_Amount_in_USD': data_list[4],
                'Purpose': data_list[5],
                'Maturity_Period': data_list[6],
                'Lender_Category': data_list[7] if data_list[7] else None,
                'Month': month,  # Corrected: Assigning the correct value to the 'Month' column
                'Year': year,    # Corrected: Assigning the correct value to the 'Year' column
                'Route': route,
            }

            cursor.execute(
                f"INSERT INTO 2005_2023_1 ({', '.join(data_dict.keys())}) VALUES ({', '.join(['%s']*len(data_dict))})",
                tuple(data_dict.values())
            )

def insert_data_into_database_49(cursor, data_list, route):
            month, year = data_list[-3], data_list[-2]
            

            data_dict = {
                'ECB_FCCB': data_list[0],
                'Borrower': data_list[1],
                'Equivalent_Amount_in_USD': data_list[2],
                'Purpose': data_list[3],
                'Maturity_Period': data_list[4],
                'Month': month,  # Corrected: Assigning the correct value to the 'Month' column
                'Year': year,    # Corrected: Assigning the correct value to the 'Year' column
                'Route': route,
            }

            cursor.execute(
                f"INSERT INTO 2005_2023_1 ({', '.join(data_dict.keys())}) VALUES ({', '.join(['%s']*len(data_dict))})",
                tuple(data_dict.values())
            )

def del2(cursor):
    cursor.execute("DELETE FROM 2005_2023_1 WHERE Equivalent_Amount_in_USD='Equivalent Amount in USD ' or Borrower='Equivalent Amount in USD' or ECB_FCCB='Borrower'or Borrower='Borrower' ")














def download_new_file(remaining_months, year, month_xpath, filename_xpath,cursor):
    # Your existing code here
    

    download_base_directory = fr"C:\Users\mohan.7482\Desktop\EXTERNAL COMMERCIL BORROWING\incremental"
    download_directory = os.path.join(download_base_directory, str(year))

    # Check if the year folder exists, create it if not
    if not os.path.exists(download_directory):
        os.makedirs(download_directory)

    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_directory,
        "download.default_content_setting_value": 2,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True
    }
    chrome_options.add_experimental_option("prefs", prefs)

    try:
        browser = webdriver.Chrome(options=chrome_options)

        browser.get("https://rbi.org.in/Scripts/ECBView.aspx")
        time.sleep(5)
        try:
            next_xpath = browser.find_element(By.XPATH, f'//*[@id="{year}"]')
            next_xpath.click()
            time.sleep(5)
        except Exception as e:
            print(year,"is not updated in the website")
            # traceback.print_exc()
            sys.exit("Error occurred. Exiting the program.")

            
             

        for month in remaining_months:
            print(month)
            xpath = month_xpath[f"{month}"]
            try:
                original_file_name = browser.find_element(By.XPATH, xpath)
            except Exception as e:
                print(month,"File is not in website")
                continue
            original_file_name = browser.find_element(By.XPATH, xpath).get_attribute("href").split("/")[-1]
            original_file_name = original_file_name.replace("%20", " ")
            print("original file name as in the website is ", original_file_name)


            downloaded_files = os.listdir(download_directory)
            print(f"the files which are in the {year} folder", downloaded_files)

            rename_xpath = browser.find_element(By.XPATH, filename_xpath[f"{month}"])
         

            original_extension=os.path.splitext(original_file_name)[1]
            print("original_extension",original_extension)


            
            new_file_name = rename_xpath.text.split("of")[-1].strip()
            print(new_file_name)

            
            new_file_name_ex=new_file_name + original_extension
            print(new_file_name_ex)
            file_name= new_file_name_ex
            if file_name in downloaded_files:
                print("no new file found for this",month,year)
                print( file_name,"the file is already exists in the folder and also inserted into the database")

            else:
                try:
                    current_month_xpath=browser.find_element(By.XPATH,xpath)
                    current_month_xpath.click()
                    time.sleep(15)
                    current_file_name = os.path.join(download_directory, original_file_name)
                    

                    if original_extension == ".xls" or original_extension == ".XLS":
                        rename_file_path = os.path.join(download_directory, file_name)
                        time.sleep(15)
                        os.rename(current_file_name, rename_file_path )
                        print(f"{year}/{file_name} - downloaded...")
                        file_path = os.path.join(download_directory,rename_file_path)
                        try:    
                            extract_and_insert_data(file_path,cursor)
                            conn.commit()
                            print(file_name,"is successfully inserted into the database")
                        except Exception as e:
                            traceback.print_exc()
                            continue
                     

                    elif original_extension == ".xlsx" or original_extension == ".XLSX" :
                        rename_file_path = os.path.join(download_directory, file_name)
                        time.sleep(15)
                        os.rename(current_file_name, rename_file_path)
                        print(f"{year}/{file_name} - downloaded...")
                        file_path = os.path.join(download_directory,rename_file_path)
                        try:    
                            extract_and_insert_data(file_path,cursor)
                            conn.commit()
                            print(file_name,"is successfully inserted into the datbase")
                        except Exception as e:
                            traceback.print_exc()
                            continue
                       
                except Exception as e:
                    traceback.print_exc()
    except Exception as e:
        traceback.print_exc()

    pass


def extract_data_from_db(year,cursor):
    remaining_months = []
    # Your existing code here
    column_name = 'year'
    desired_value = year

    query = f"SELECT * FROM 2005_2023_1 WHERE {column_name} = %s"
    cursor.execute(query, (desired_value,))

    # Fetch rows based on the condition
    selected_rows = cursor.fetchall()

    # Create a list to store the 10th values without duplicates
    months_in_database = set()
    year_database = set()

    # Extract the 10th values from each selected row and add to the list
    for row in selected_rows:
        if len(row) > 9:  
            months_in_database.add(row[8])
            year_database.add(row[9])


    # Print the unique 10th values
    print("These Months data are already in the database:", months_in_database)
    print("year",year_database)

    for a_month in actual_months:
        if a_month in months_in_database:
            print(a_month,"These month data is already inserted into the database")
        else:
            remaining_months.append(a_month)


    print(remaining_months,"These months are not in the database")
    if len(remaining_months) > 0:
        download_new_file(remaining_months,year, month_xpath, filename_xpath,cursor)
    else:
        print("no months want to be insert into the database")
        print(remaining_months,"All the months are inserted for this year",year)

    pass






cursor.execute("SELECT * FROM 2005_2023_1")

# Fetch all rows
all_rows = cursor.fetchall()

# Print the data
# for row in all_rows:
#     print(row)



# find current year and previous year

years = []


def get_current_year():
    current_year = datetime.now().year
    years.append(current_year)
    return current_year


current_year = get_current_year()
previous_year = current_year - 1
years.append(previous_year)
print("The current year is:",current_year)
print("The previous year is:",previous_year)


for year in years:
    extract_data_from_db(year,cursor)
 

# Close the cursor and connection

del2(cursor)
conn.commit()
cursor.close()
conn.close()
